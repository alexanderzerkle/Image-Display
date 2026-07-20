#!/usr/bin/env python3
"""Render a published Google Sheet XLSX workbook as an 800x480 PNG."""

from __future__ import annotations

import io
import os
import textwrap
import urllib.request
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Color, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


SHEET_XLSX_URL = os.environ.get(
    "SHEET_XLSX_URL",
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vQDVv0Nt9doPsH3SQgnFUHbvcs8i_s_iCw727ZP2sBX1Ty5RNjWRUCswVSxTa8qVLD5XAziLgN4IC60/"
    "pub?output=xlsx",
)

OUTPUT_DIR = Path(os.environ.get("OUTPUT_DIR", "dist"))
WORKSHEET_NAME = os.environ.get("WORKSHEET_NAME", "").strip()

MAX_ROWS = int(os.environ.get("MAX_ROWS", "40"))
MAX_COLUMNS = int(os.environ.get("MAX_COLUMNS", "20"))

WIDTH = 800
HEIGHT = 480
MARGIN = 10

DEFAULT_ROW_HEIGHT = 22.0
DEFAULT_COLUMN_WIDTH = 10.0

MIN_FONT_SIZE = 9
MAX_FONT_SIZE = 32


def download_xlsx(url: str) -> bytes:
    """Download the published workbook and verify that it looks like an XLSX file."""

    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "github-pages-sheet-renderer/2.0",
        },
    )

    with urllib.request.urlopen(request, timeout=45) as response:
        data = response.read()
        content_type = response.headers.get("Content-Type", "")

    # XLSX files are ZIP containers and normally begin with PK.
    if not data.startswith(b"PK"):
        preview = data[:200].decode("utf-8", errors="replace")

        raise RuntimeError(
            "The spreadsheet URL did not return an XLSX file. "
            f"Content-Type was {content_type!r}. "
            f"Response began with: {preview!r}"
        )

    return data


def load_font(
    size: int,
    bold: bool = False,
    italic: bool = False,
) -> ImageFont.ImageFont:
    """Load a font normally available on GitHub's Ubuntu runners."""

    size = max(MIN_FONT_SIZE, min(MAX_FONT_SIZE, int(round(size))))

    if bold and italic:
        names = [
            "DejaVuSans-BoldOblique.ttf",
            "LiberationSans-BoldItalic.ttf",
        ]
    elif bold:
        names = [
            "DejaVuSans-Bold.ttf",
            "LiberationSans-Bold.ttf",
        ]
    elif italic:
        names = [
            "DejaVuSans-Oblique.ttf",
            "LiberationSans-Italic.ttf",
        ]
    else:
        names = [
            "DejaVuSans.ttf",
            "LiberationSans-Regular.ttf",
        ]

    roots = [
        Path("/usr/share/fonts/truetype/dejavu"),
        Path("/usr/share/fonts/truetype/liberation2"),
    ]

    for root in roots:
        for name in names:
            path = root / name

            if path.exists():
                return ImageFont.truetype(str(path), size=size)

    return ImageFont.load_default()


def color_to_rgb(
    color: Color | None,
    fallback: tuple[int, int, int],
) -> tuple[int, int, int]:
    """Convert a basic openpyxl color to an RGB tuple."""

    if color is None:
        return fallback

    if color.type == "rgb" and color.rgb:
        value = color.rgb[-6:]

        try:
            return tuple(
                int(value[index : index + 2], 16)
                for index in (0, 2, 4)
            )
        except ValueError:
            return fallback

    if color.type == "indexed" and color.indexed is not None:
        palette = {
            0: (0, 0, 0),
            1: (255, 255, 255),
            2: (255, 0, 0),
            3: (0, 255, 0),
            4: (0, 0, 255),
            5: (255, 255, 0),
            6: (255, 0, 255),
            7: (0, 255, 255),
            8: (0, 0, 0),
            9: (255, 255, 255),
        }

        return palette.get(int(color.indexed), fallback)

    # Theme colors require workbook-theme calculations.
    # Fall back safely rather than failing.
    return fallback


def fill_rgb(fill: PatternFill | None) -> tuple[int, int, int]:
    """Return the cell's solid fill color."""

    if fill is None:
        return (255, 255, 255)

    if fill.fill_type != "solid":
        return (255, 255, 255)

    return color_to_rgb(
        fill.fgColor,
        (255, 255, 255),
    )


def luminance(rgb: tuple[int, int, int]) -> float:
    r, g, b = rgb

    return (
        0.2126 * r
        + 0.7152 * g
        + 0.0722 * b
    )


def text_color(
    cell,
    background: tuple[int, int, int],
) -> tuple[int, int, int]:
    """Use the workbook's font color, or choose readable black/white."""

    sentinel = (-1, -1, -1)

    explicit = color_to_rgb(
        getattr(cell.font, "color", None),
        sentinel,
    )

    if explicit != sentinel:
        return explicit

    if luminance(background) < 110:
        return (255, 255, 255)

    return (0, 0, 0)


def row_height_pixels(sheet, row: int) -> float:
    """Convert Excel row height in points to approximate pixels."""

    points = (
        sheet.row_dimensions[row].height
        or DEFAULT_ROW_HEIGHT
    )

    return max(
        10.0,
        float(points) * 96.0 / 72.0,
    )


def column_width_pixels(sheet, column: int) -> float:
    """Convert Excel column width to approximate pixels."""

    letter = get_column_letter(column)

    width = (
        sheet.column_dimensions[letter].width
        or DEFAULT_COLUMN_WIDTH
    )

    return max(
        18.0,
        float(width) * 7.0 + 5.0,
    )


def used_bounds(sheet) -> tuple[int, int]:
    """Find the useful sheet area within the configured maximums."""

    max_row = min(sheet.max_row, MAX_ROWS)
    max_col = min(sheet.max_column, MAX_COLUMNS)

    while max_row > 1:
        row_is_empty = all(
            sheet.cell(max_row, column).value in (None, "")
            for column in range(1, max_col + 1)
        )

        if not row_is_empty:
            break

        max_row -= 1

    while max_col > 1:
        column_is_empty = all(
            sheet.cell(row, max_col).value in (None, "")
            for row in range(1, max_row + 1)
        )

        if not column_is_empty:
            break

        max_col -= 1

    return max_row, max_col


def natural_edges(
    sizes: Iterable[float],
    start: int,
    scale: float,
) -> list[int]:
    """Convert worksheet sizes into pixel edges using one uniform scale."""

    edges = [start]
    running = float(start)

    for value in sizes:
        running += value * scale
        edges.append(round(running))

    return edges


def worksheet_scales(
    row_sizes: list[float],
    col_sizes: list[float],
) -> tuple[float, float]:
    """Scale columns and rows separately to use the 800x480 canvas.

    Google Sheets itself displays column widths and row heights using different
    screen conversions.  A single uniform scale made wide sheets force the
    rows and fonts to become far too small.  Independent axes preserve the
    spreadsheet-like density while fitting the useful range to the image.
    """

    natural_width = sum(col_sizes) or 1.0
    natural_height = sum(row_sizes) or 1.0
    available_width = WIDTH - 2 * MARGIN
    available_height = HEIGHT - 2 * MARGIN

    return (
        available_width / natural_width,
        available_height / natural_height,
    )


def cell_has_visual_content(cell) -> bool:
    """Return True when a cell should stop text overflowing through it."""

    if cell.value not in (None, ""):
        return True

    fill = getattr(cell, "fill", None)
    if fill is not None and getattr(fill, "fill_type", None):
        return True

    border = getattr(cell, "border", None)
    if border is not None:
        for name in ("left", "right", "top", "bottom"):
            side = getattr(border, name, None)
            if side is not None and getattr(side, "style", None):
                return True

    return False

def merged_anchor_map(
    sheet,
    max_row: int,
    max_col: int,
) -> tuple[
    dict[tuple[int, int], tuple[int, int]],
    dict[tuple[int, int], tuple[int, int, int, int]],
]:
    """Record merged-cell anchors and the cells covered by each merge."""

    covered: dict[
        tuple[int, int],
        tuple[int, int],
    ] = {}

    spans: dict[
        tuple[int, int],
        tuple[int, int, int, int],
    ] = {}

    for merged in sheet.merged_cells.ranges:
        min_col, min_row, end_col, end_row = merged.bounds

        if min_row > max_row or min_col > max_col:
            continue

        end_row = min(end_row, max_row)
        end_col = min(end_col, max_col)

        anchor = (min_row, min_col)

        spans[anchor] = (
            min_row,
            min_col,
            end_row,
            end_col,
        )

        for row in range(min_row, end_row + 1):
            for col in range(min_col, end_col + 1):
                if (row, col) != anchor:
                    covered[(row, col)] = anchor

    return covered, spans


def fit_single_line(
    draw: ImageDraw.ImageDraw,
    text: str,
    font: ImageFont.ImageFont,
    width: int,
) -> str:
    """Fit one line inside the available width, adding an ellipsis if needed."""

    text = text.replace("\r", " ").replace("\n", " ").strip()

    if not text:
        return ""

    if draw.textlength(text, font=font) <= width:
        return text

    ellipsis = "..."
    shortened = text

    while (
        shortened
        and draw.textlength(
            shortened + ellipsis,
            font=font,
        ) > width
    ):
        shortened = shortened[:-1]

    return shortened.rstrip() + ellipsis if shortened else ellipsis


def wrap_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    font: ImageFont.ImageFont,
    width: int,
    max_lines: int,
    allow_wrap: bool,
) -> list[str]:
    """
    Fit text into a cell.

    Explicit line breaks are always preserved. Automatic word wrapping is
    performed only when Excel's Wrap Text setting is enabled.
    """

    text = str(text).replace("\r\n", "\n").replace("\r", "\n").strip()

    if not text:
        return []

    paragraphs = text.split("\n")
    lines: list[str] = []

    for paragraph in paragraphs:
        if not allow_wrap:
            lines.append(
                fit_single_line(
                    draw,
                    paragraph,
                    font,
                    width,
                )
            )
            continue

        words = paragraph.split()

        if not words:
            lines.append("")
            continue

        current = ""

        for word in words:
            candidate = word if not current else f"{current} {word}"

            if draw.textlength(candidate, font=font) <= width:
                current = candidate
                continue

            if current:
                lines.append(current)
                current = ""

            # Break a single word that is wider than the cell.
            remaining = word

            while remaining:
                chunk = ""

                while (
                    remaining
                    and draw.textlength(
                        chunk + remaining[0],
                        font=font,
                    ) <= width
                ):
                    chunk += remaining[0]
                    remaining = remaining[1:]

                if not chunk:
                    chunk = remaining[0]
                    remaining = remaining[1:]

                if remaining:
                    lines.append(chunk)
                else:
                    current = chunk

        if current:
            lines.append(current)

    if len(lines) <= max_lines:
        return lines

    lines = lines[:max_lines]

    ellipsis = "..."
    last = lines[-1]

    while (
        last
        and draw.textlength(
            last + ellipsis,
            font=font,
        ) > width
    ):
        last = last[:-1]

    lines[-1] = last.rstrip() + ellipsis

    return lines


def draw_checkbox(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    checked: bool,
    color: tuple[int, int, int],
) -> None:
    """Draw a checked or unchecked box for Boolean cells."""

    left, top, right, bottom = box

    size = max(
        6,
        min(
            18,
            right - left - 4,
            bottom - top - 4,
        ),
    )

    x = left + (right - left - size) // 2
    y = top + (bottom - top - size) // 2

    draw.rectangle(
        (
            x,
            y,
            x + size,
            y + size,
        ),
        outline=color,
        width=2,
    )

    if checked:
        draw.line(
            (
                x + round(size * 0.20),
                y + round(size * 0.55),
                x + round(size * 0.43),
                y + round(size * 0.78),
                x + round(size * 0.82),
                y + round(size * 0.25),
            ),
            fill=color,
            width=max(2, size // 7),
            joint="curve",
        )


def draw_simple_border(
    draw: ImageDraw.ImageDraw,
    border: Border | None,
    box: tuple[int, int, int, int],
) -> None:
    """Draw simple one-pixel versions of explicit Excel borders."""

    left, top, right, bottom = box

    sides = {
        "top": (
            getattr(border, "top", None)
            if border is not None
            else None
        ),
        "right": (
            getattr(border, "right", None)
            if border is not None
            else None
        ),
        "bottom": (
            getattr(border, "bottom", None)
            if border is not None
            else None
        ),
        "left": (
            getattr(border, "left", None)
            if border is not None
            else None
        ),
    }

    def visible(side) -> bool:
        return (
            side is not None
            and getattr(side, "style", None) is not None
        )

    if visible(sides["top"]):
        draw.line(
            (left, top, right, top),
            fill=(0, 0, 0),
            width=1,
        )

    if visible(sides["right"]):
        draw.line(
            (right, top, right, bottom),
            fill=(0, 0, 0),
            width=1,
        )

    if visible(sides["bottom"]):
        draw.line(
            (left, bottom, right, bottom),
            fill=(0, 0, 0),
            width=1,
        )

    if visible(sides["left"]):
        draw.line(
            (left, top, left, bottom),
            fill=(0, 0, 0),
            width=1,
        )


def draw_text_underline(
    draw: ImageDraw.ImageDraw,
    font,
    underline_style: str | None,
    text_x: float,
    text_y: float,
    text_width: float,
    line_height: int,
    cell_left: int,
    cell_right: int,
    padding: int,
    color: tuple[int, int, int],
) -> None:
    """Draw Excel-style font underlines beneath a rendered line of text."""

    if not underline_style:
        return

    style = str(underline_style)

    is_double = style in {
        "double",
        "doubleAccounting",
    }

    is_accounting = style in {
        "singleAccounting",
        "doubleAccounting",
    }

    if is_accounting:
        start_x = cell_left + padding
        end_x = cell_right - padding
    else:
        start_x = round(text_x)
        end_x = round(text_x + text_width)

    # Place the underline near the font baseline rather than at the very
    # bottom of the cell.
    ascent, _descent = font.getmetrics()
    underline_y = round(text_y + min(ascent + 1, line_height - 2))

    draw.line(
        (
            start_x,
            underline_y,
            end_x,
            underline_y,
        ),
        fill=color,
        width=1,
    )

    if is_double:
        draw.line(
            (
                start_x,
                underline_y + 2,
                end_x,
                underline_y + 2,
            ),
            fill=color,
            width=1,
        )

def draw_cell_text(
    draw: ImageDraw.ImageDraw,
    sheet,
    cell,
    box: tuple[int, int, int, int],
    background: tuple[int, int, int],
    col_edges: list[int],
    max_col: int,
    scale: float,
) -> None:
    """Draw a cell's text or Boolean checkbox."""

    value = cell.value

    if value is None or value == "":
        return

    if isinstance(value, bool):
        draw_checkbox(
            draw,
            box,
            value,
            text_color(cell, background),
        )
        return

    left, top, right, bottom = box
    padding = max(2, round(4 * scale))

    alignment: Alignment = cell.alignment or Alignment()
    raw_text = str(value)
    allow_wrap = bool(alignment.wrap_text) or "\n" in raw_text or "\r" in raw_text

    # Google Sheets lets unwrapped text flow into adjacent empty cells.
    text_right = right
    if not allow_wrap and (alignment.horizontal or "left") in {"left", "general"}:
        for next_col in range(cell.column + 1, max_col + 1):
            next_cell = sheet.cell(cell.row, next_col)
            if cell_has_visual_content(next_cell):
                break
            text_right = col_edges[next_col]

    available_width = max(1, text_right - left - padding * 2)
    available_height = max(1, bottom - top - padding * 2)

    size = float(cell.font.sz or 11) * scale
    font = load_font(
        size,
        bool(cell.font.bold),
        bool(cell.font.italic),
    )

    color = text_color(cell, background)
    line_bbox = draw.textbbox((0, 0), "Ag", font=font)
    line_height = max(1, line_bbox[3] - line_bbox[1] + max(1, round(2 * scale)))
    max_lines = max(1, available_height // line_height)

    lines = wrap_text(
        draw,
        raw_text,
        font,
        available_width,
        max_lines,
        allow_wrap,
    )

    if not lines:
        return

    block_height = len(lines) * line_height
    vertical = alignment.vertical or "center"

    if vertical == "top":
        y = top + padding
    elif vertical == "bottom":
        y = bottom - padding - block_height
    else:
        y = top + (bottom - top - block_height) // 2

    horizontal = alignment.horizontal or "left"
    underline_style = getattr(cell.font, "underline", None)

    for line in lines:
        line_width = draw.textlength(line, font=font)

        if horizontal in {"center", "centerContinuous"}:
            x = left + (right - left - line_width) / 2
        elif horizontal == "right":
            x = right - padding - line_width
        else:
            x = left + padding

        draw.text((round(x), round(y)), line, font=font, fill=color)

        draw_text_underline(
            draw=draw,
            font=font,
            underline_style=underline_style,
            text_x=x,
            text_y=y,
            text_width=line_width,
            line_height=line_height,
            cell_left=left,
            cell_right=right,
            padding=padding,
            color=color,
        )

        y += line_height

def render_png(
    workbook_bytes: bytes,
    output_path: Path,
) -> None:
    """Load the workbook and render its selected worksheet."""

    workbook = load_workbook(
        io.BytesIO(workbook_bytes),
        data_only=True,
    )

    if WORKSHEET_NAME:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise RuntimeError(
                f"Worksheet {WORKSHEET_NAME!r} was not found. "
                f"Available sheets: {workbook.sheetnames}"
            )

        sheet = workbook[WORKSHEET_NAME]
    else:
        sheet = workbook.active

    max_row, max_col = used_bounds(sheet)

    row_sizes = [
        row_height_pixels(sheet, row)
        for row in range(1, max_row + 1)
    ]
    col_sizes = [
        column_width_pixels(sheet, col)
        for col in range(1, max_col + 1)
    ]

    x_scale, y_scale = worksheet_scales(row_sizes, col_sizes)
    row_edges = natural_edges(row_sizes, MARGIN, y_scale)
    col_edges = natural_edges(col_sizes, MARGIN, x_scale)
    font_scale = y_scale

    covered, spans = merged_anchor_map(
        sheet,
        max_row,
        max_col,
    )

    image = Image.new(
        "RGB",
        (WIDTH, HEIGHT),
        (255, 255, 255),
    )

    draw = ImageDraw.Draw(image)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if (row, col) in covered:
                continue

            cell = sheet.cell(row, col)

            if isinstance(cell, MergedCell):
                continue

            start_row, start_col, end_row, end_col = spans.get(
                (row, col),
                (row, col, row, col),
            )

            box = (
                col_edges[start_col - 1],
                row_edges[start_row - 1],
                col_edges[end_col],
                row_edges[end_row],
            )

            background = fill_rgb(cell.fill)

            draw.rectangle(
                box,
                fill=background,
            )

            draw_simple_border(
                draw,
                cell.border,
                box,
            )

            draw_cell_text(
                draw,
                sheet,
                cell,
                box,
                background,
                col_edges,
                max_col,
                font_scale,
            )


    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if (row, col) in covered:
                continue

            cell = sheet.cell(row, col)

            if isinstance(cell, MergedCell):
                continue

            start_row, start_col, end_row, end_col = spans.get(
                (row, col),
                (row, col, row, col),
            )

            box = (
                col_edges[start_col - 1],
                row_edges[start_row - 1],
                col_edges[end_col],
                row_edges[end_row],
            )

            draw_simple_border(
                draw,
                cell.border,
                box,
            )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    image.save(
        output_path,
        format="PNG",
        optimize=True,
    )


def write_index(output_dir: Path) -> None:
    """Create the simple GitHub Pages wrapper page."""

    html = """<!doctype html>
    <html lang="en">
    <head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1">
      <title>Google Sheet PNG</title>
      <style>
        html, body {
          min-height: 100%;
        }

        body {
          margin: 0;
          display: grid;
          place-items: center;
          background: #eee;
          font-family: sans-serif;
        }

        main {
          max-width: 840px;
          padding: 20px;
          text-align: center;
        }

        img {
          display: block;
          width: min(800px, 100%);
          height: auto;
          border: 1px solid #000;
        }

        p {
          margin-bottom: 0;
        }
      </style>
    </head>
    <body>
      <main>
        <img
          src="sheet.png"
          width="800"
          height="480"
          alt="Spreadsheet rendered as an image"
        >
        <p>
          <a href="sheet.png">Open the PNG directly</a>
        </p>
      </main>
    </body>
    </html>
    """

    (output_dir / "index.html").write_text(
        textwrap.dedent(html),
        encoding="utf-8",
    )

    (output_dir / ".nojekyll").write_text(
        "",
        encoding="utf-8",
    )


def main() -> None:
    workbook_bytes = download_xlsx(
        SHEET_XLSX_URL,
    )

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    render_png(
        workbook_bytes,
        OUTPUT_DIR / "sheet.png",
    )

    write_index(
        OUTPUT_DIR,
    )

    print(
        f"Created {OUTPUT_DIR / 'sheet.png'} "
        f"({WIDTH}x{HEIGHT})"
    )


if __name__ == "__main__":
    main()
